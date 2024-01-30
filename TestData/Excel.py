import openpyxl
book = openpyxl.load_workbook("E:\\SLENIMUM\\Excel files\\pythonselenium.xlsx")
sheet = book.active
Dict ={}
cell = sheet.cell(row=1, column=2)
#print(cell.value)
sheet.cell(row=2, column=2).value = "avnish"
# print(sheet.cell(row=2, column=2).value)
# print(sheet.max_row)
# print(sheet.max_column)
# print(sheet['B1'].value)
for i in range(2, sheet.max_row+1):  # to get the rows
    if sheet.cell(row=i, column=1).value == "Testcase2":

        for j in range(2, sheet.max_column+1):  # To get the Columns
            # Dict["lastname"]= "kumar"
            Dict[sheet.cell(row=i, column=j).value] = sheet.cell(row=i, column=j).value



print(sheet['D8'].value)
print(Dict)
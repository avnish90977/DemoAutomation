import openpyxl


class HomePageData:
    test_HomePage_Data = [{"FirstName": "Gautam", "Email": "yadavavnishraj@gmail.com", "Password": "Password",
                           "Gender": "Male"},
                          {"FirstName": "Sneha", "Email": "snehagupta@gmail.com", "Password": "Password234",
                           "Gender": "Female"}]



    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("E:\\SLENIMUM\\Excel files\\pythonselenium.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row +1):  # to get the rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # To get the Columns
                    # Dict["lastname"]= "kumar"
                    Dict[sheet.cell(row=i, column=j).value] =sheet.cell(row=i, column=j).value
        return[Dict]
